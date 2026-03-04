"""Report builders and exporters for Code Learning Analyzer."""

from __future__ import annotations

from pathlib import Path

from analyzer_core import (
    build_advanced_stats_map,
    language_usage_stats,
    learning_tips,
    metric_explanations,
    summarize,
)
from analyzer_models import FileStats, LONG_LINE_LIMIT


def build_report(
    path: Path,
    results: list[FileStats],
    include_detailed_stats: bool = False,
    include_very_detailed_stats: bool = False,
    explain_metrics: bool = False,
) -> str:
    """Build a comprehensive text report from analysis results."""
    summary = summarize(results)
    # Zaawansowane metryki liczymy tylko wtedy, gdy uzytkownik faktycznie o nie poprosil.
    advanced_stats_map = build_advanced_stats_map(results) if (include_detailed_stats or include_very_detailed_stats) else {}

    lines: list[str] = []
    lines.append("=== Prosty Raport Kodu v2 ===")
    lines.append(f"Sciezka: {path}")
    lines.append(f"Liczba plikow: {summary.files_count}")
    lines.append(f"Wszystkie linie: {summary.total_lines}")
    lines.append(f"Linie kodu: {summary.code_lines}")
    lines.append(f"Linie puste: {summary.empty_lines}")
    lines.append(f"Linie komentarzy: {summary.comment_lines}")
    lines.append(f"Wykryte funkcje: {summary.function_count}")
    lines.append(f"Linie > {LONG_LINE_LIMIT} znakow: {summary.long_lines}")
    lines.append(f"Liczba TODO: {summary.todo_count}")

    # Sekcja edukacyjna jest opcjonalna, bo bywa zbyt dluga przy szybkich raportach.
    if explain_metrics:
        lines.append("")
        lines.append("=== Jak czytac metryki ===")
        for idx, explanation in enumerate(metric_explanations(), start=1):
            lines.append(f"{idx}. {explanation}")

    lines.append("")
    lines.append("=== Najwieksze pliki (top 5) ===")
    biggest = sorted(results, key=lambda item: item.total_lines, reverse=True)[:5]
    if not biggest:
        lines.append("Brak plikow do analizy.")
    else:
        for idx, file_stats in enumerate(biggest, start=1):
            lines.append(f"{idx}. {file_stats.file} ({file_stats.total_lines} linii)")

    language_rows = language_usage_stats(results)
    lines.append("")
    lines.append("=== Uzycie jezykow ===")
    for row in language_rows:
        lines.append(
            f"- {row['language']}: {row['files']} plik(ow), {row['code_lines']} linii kodu ({row['code_share']:.1%})"
        )

    # Szczegoly metryk pojawiaja sie warunkowo, zeby raport podstawowy pozostal krotki i czytelny.
    if include_detailed_stats or include_very_detailed_stats:
        lines.append("")
        lines.append("=== Statystyki zaawansowane ===")
        for file_stats in results:
            advanced = advanced_stats_map[file_stats.file]
            complexity_per_100 = (advanced.cyclomatic_complexity / max(file_stats.code_lines, 1)) * 100
            lines.append(f"\n{file_stats.file}:")
            lines.append(f"  - Zlozonosc cyklomatyczna: {advanced.cyclomatic_complexity}")
            lines.append(f"  - Zlozonosc / 100 linii kodu: {complexity_per_100:.1f}")
            lines.append(f"  - Gestosc kodu: {advanced.code_density:.2%}")
            lines.append(f"  - Wspolczynnik duplikacji: {advanced.duplication_score:.2%}")
            lines.append(f"  - Importy: {advanced.import_count}")
            lines.append(f"  - Klasy: {advanced.class_count}")
            lines.append(f"  - Petle: {advanced.loop_count}")
            lines.append(f"  - If-y: {advanced.if_count}")
            lines.append(f"  - Try-catch: {advanced.try_catch_count}")

    # Najbardziej szczegolowy tryb dokleja konkretne linie problematyczne do szybkiej poprawy.
    if include_very_detailed_stats:
        lines.append("")
        lines.append("=== Szczegóły problemów ===")
        for file_stats in results:
            if file_stats.long_lines_details:
                lines.append(f"\n{file_stats.file} - Długie linie:")
                for line_num, content in file_stats.long_lines_details[:3]:
                    lines.append(f"  {line_num}: {content[:80]}...")

            if file_stats.todo_details:
                lines.append(f"\n{file_stats.file} - TODO:")
                for line_num, content in file_stats.todo_details:
                    lines.append(f"  {line_num}: {content}")

            if file_stats.function_details:
                lines.append(f"\n{file_stats.file} - Funkcje:")
                for line_num, content in file_stats.function_details[:5]:
                    lines.append(f"  {line_num}: {content}")

    lines.append("")
    lines.append("=== Co cwiczyc dalej ===")
    for idx, tip in enumerate(learning_tips(summary, results), start=1):
        lines.append(f"{idx}. {tip}")

    return "\n".join(lines)


def create_visualization_report(path: Path, results: list[FileStats], out_path: Path) -> None:
    """Create a comprehensive PDF visualization report with charts."""
    try:
        import matplotlib
        # Backend "Agg" pozwala generowac wykresy bez otwierania okna GUI (np. na serwerze/CI).
        matplotlib.use("Agg", force=True)
        import matplotlib.pyplot as plt
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError(
            "Brak bibliotek matplotlib lub reportlab. Zainstaluj: python3 -m pip install matplotlib reportlab"
        ) from exc

    summary = summarize(results)
    biggest = sorted(results, key=lambda item: item.total_lines, reverse=True)[:5]
    tips = learning_tips(summary, results)
    advanced_stats_map = build_advanced_stats_map(results)

    plt.style.use("default")
    plt.rcParams["figure.dpi"] = 150
    plt.rcParams["savefig.dpi"] = 300
    plt.rcParams["font.size"] = 10

    import os
    import tempfile

    # Wykresy skladamy najpierw do PNG w katalogu tymczasowym, potem osadzamy je w PDF.
    temp_dir = tempfile.mkdtemp()

    try:
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(11, 8.5))
        fig.suptitle(f"Analiza Kodu: {path.name}", fontsize=16, fontweight="bold")

        labels = ["Kod", "Komentarze", "Puste"]
        sizes = [summary.code_lines, summary.comment_lines, summary.empty_lines]
        colors_pie = ["#2E86AB", "#F18F01", "#C73E1D"]
        ax1.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors_pie, startangle=90)
        ax1.set_title("Podział linii kodu", fontsize=12, fontweight="bold")

        basic_values = [summary.files_count, summary.code_lines, summary.function_count]
        ax2.bar(["Pliki", "Linie kodu", "Funkcje"], basic_values, color=["#4ECDC4", "#FF6B6B", "#45B7D1"])
        ax2.set_title("Podstawowe statystyki", fontsize=12, fontweight="bold")
        ax2.set_ylabel("Liczba")
        for idx, value in enumerate(basic_values):
            ax2.text(idx, value + max(basic_values) * 0.05, str(value), ha="center", fontweight="bold")

        issue_values = [summary.long_lines, summary.todo_count]
        ax3.bar(["Długie linie", "TODO"], issue_values, color=["#FF4757", "#2ED573"])
        ax3.set_title("Problemy do poprawy", fontsize=12, fontweight="bold")
        ax3.set_ylabel("Liczba")
        for idx, value in enumerate(issue_values):
            ax3.text(idx, value + max(issue_values) * 0.05, str(value), ha="center", fontweight="bold")

        complexity_data = []
        file_names = []
        for file_stats in results[:5]:
            advanced = advanced_stats_map[file_stats.file]
            complexity_data.append(advanced.cyclomatic_complexity)
            short_name = Path(file_stats.file).name
            file_names.append(short_name[:15] + "..." if len(short_name) > 15 else short_name)

        bars = ax4.bar(range(len(complexity_data)), complexity_data, color="#A555EC")
        ax4.set_title("Złożoność cyklomatyczna (top 5)", fontsize=12, fontweight="bold")
        ax4.set_ylabel("Złożoność")
        ax4.set_xticks(range(len(complexity_data)))
        ax4.set_xticklabels(file_names, rotation=45, ha="right")
        for bar in bars:
            height = bar.get_height()
            ax4.text(bar.get_x() + bar.get_width() / 2.0, height + 0.5, f"{int(height)}", ha="center", va="bottom", fontweight="bold")

        plt.tight_layout()
        page1_path = os.path.join(temp_dir, "page1.png")
        plt.savefig(page1_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 8.5))
        fig.suptitle(f"Porównanie plików - {path.name}", fontsize=16, fontweight="bold")

        file_names = [Path(item.file).name for item in biggest]
        file_lines = [item.total_lines for item in biggest]
        bars1 = ax1.bar(range(len(file_names)), file_lines, color="#FF9F43")
        ax1.set_title("Największe pliki", fontsize=12, fontweight="bold")
        ax1.set_ylabel("Liczba linii")
        ax1.set_xticks(range(len(file_names)))
        ax1.set_xticklabels(file_names, rotation=45, ha="right")
        for bar in bars1:
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width() / 2.0, height + max(file_lines) * 0.02, f"{height}", ha="center", va="bottom", fontweight="bold")

        function_counts = [item.function_count for item in biggest]
        bars2 = ax2.bar(range(len(file_names)), function_counts, color="#54A0FF")
        ax2.set_title("Funkcje w plikach", fontsize=12, fontweight="bold")
        ax2.set_ylabel("Liczba funkcji")
        ax2.set_xticks(range(len(file_names)))
        ax2.set_xticklabels(file_names, rotation=45, ha="right")
        for bar in bars2:
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width() / 2.0, height + max(function_counts) * 0.02, f"{height}", ha="center", va="bottom", fontweight="bold")

        plt.tight_layout()
        page2_path = os.path.join(temp_dir, "page2.png")
        plt.savefig(page2_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 8.5))
        fig.suptitle(f"Analiza problemów - {path.name}", fontsize=16, fontweight="bold")

        files_with_issues = [item for item in results if item.long_lines > 0 or item.todo_count > 0][:5]
        if files_with_issues:
            issue_names = [Path(item.file).name for item in files_with_issues]
            long_lines_data = [item.long_lines for item in files_with_issues]
            todo_data = [item.todo_count for item in files_with_issues]
            x = range(len(issue_names))
            width = 0.35

            bars1 = ax1.bar([i - width / 2 for i in x], long_lines_data, width, label="Długie linie", color="#FF6B6B")
            bars2 = ax1.bar([i + width / 2 for i in x], todo_data, width, label="TODO", color="#4ECDC4")
            ax1.set_title("Problemy w plikach", fontsize=12, fontweight="bold")
            ax1.set_ylabel("Liczba")
            ax1.set_xticks(x)
            ax1.set_xticklabels(issue_names, rotation=45, ha="right")
            ax1.legend()

            for bars in [bars1, bars2]:
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax1.text(
                            bar.get_x() + bar.get_width() / 2.0,
                            height + max(max(long_lines_data, default=0), max(todo_data, default=0)) * 0.02,
                            f"{int(height)}",
                            ha="center",
                            va="bottom",
                            fontweight="bold",
                        )
        else:
            ax1.text(0.5, 0.5, "Brak problemów do wyświetlenia", ha="center", va="center", transform=ax1.transAxes)
            ax1.set_title("Problemy w plikach", fontsize=12, fontweight="bold")

        if results:
            avg_complexity = sum(advanced_stats_map[item.file].cyclomatic_complexity for item in results) / len(results)
            avg_density = sum(advanced_stats_map[item.file].code_density for item in results) / len(results)
            avg_duplication = sum(advanced_stats_map[item.file].duplication_score for item in results) / len(results)
            metrics = ["Złożoność", "Gęstość", "Duplikacja"]
            values = [avg_complexity, avg_density * 100, avg_duplication * 100]
            bars = ax2.bar(metrics, values, color=["#A555EC", "#54A0FF", "#FF9F43"])
            ax2.set_title("Średnie metryki jakości", fontsize=12, fontweight="bold")
            ax2.set_ylabel("Wartość")
            for bar, value in zip(bars, values):
                ax2.text(bar.get_x() + bar.get_width() / 2.0, bar.get_height() + max(values) * 0.02, f"{value:.1f}", ha="center", va="bottom", fontweight="bold")
        else:
            ax2.text(0.5, 0.5, "Brak danych do wyświetlenia", ha="center", va="center", transform=ax2.transAxes)
            ax2.set_title("Średnie metryki jakości", fontsize=12, fontweight="bold")

        plt.tight_layout()
        page3_path = os.path.join(temp_dir, "page3.png")
        plt.savefig(page3_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, ax = plt.subplots(figsize=(11, 8.5))
        fig.suptitle(f"Porady i rekomendacje - {path.name}", fontsize=16, fontweight="bold")
        recommendations_text = "Najważniejsze rekomendacje:\n\n"
        for idx, tip in enumerate(tips[:8], 1):
            recommendations_text += f"{idx}. {tip}\n\n"

        ax.text(
            0.05,
            0.95,
            recommendations_text,
            transform=ax.transAxes,
            fontsize=11,
            verticalalignment="top",
            bbox=dict(boxstyle="round", facecolor="#F0F0F0", alpha=0.8),
        )
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")

        plt.tight_layout()
        page4_path = os.path.join(temp_dir, "page4.png")
        plt.savefig(page4_path, dpi=150, bbox_inches="tight")
        plt.close()

        # ReportLab laczy gotowe obrazy w wielostronicowy raport.
        pdf = canvas.Canvas(str(out_path), pagesize=landscape(A4))
        width, height = landscape(A4)

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(15 * mm, height - 15 * mm, f"Strona 1: Podsumowanie projektu - {path.name}")
        pdf.drawImage(page1_path, 15 * mm, 80 * mm, width - 30 * mm, 120 * mm)
        pdf.showPage()

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(15 * mm, height - 15 * mm, f"Strona 2: Porównanie plików - {path.name}")
        pdf.drawImage(page2_path, 15 * mm, 80 * mm, width - 30 * mm, 120 * mm)
        pdf.showPage()

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(15 * mm, height - 15 * mm, f"Strona 3: Analiza problemów - {path.name}")
        pdf.drawImage(page3_path, 15 * mm, 80 * mm, width - 30 * mm, 120 * mm)
        pdf.showPage()

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(15 * mm, height - 15 * mm, f"Strona 4: Porady i rekomendacje - {path.name}")
        pdf.drawImage(page4_path, 15 * mm, 80 * mm, width - 30 * mm, 120 * mm)
        pdf.showPage()

        pdf.save()
    finally:
        import shutil

        # Sprzatamy pliki tymczasowe zawsze, nawet gdy generowanie raportu sie nie powiedzie.
        shutil.rmtree(temp_dir)


def write_xlsx_report(path: Path, results: list[FileStats], out_path: Path) -> None:
    """Write analysis results to an Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "Brak biblioteki openpyxl. Zainstaluj: python3 -m pip install openpyxl"
        ) from exc

    summary = summarize(results)
    biggest = sorted(results, key=lambda item: item.total_lines, reverse=True)[:5]
    tips = learning_tips(summary, results)
    advanced_stats_map = build_advanced_stats_map(results)

    workbook = Workbook()
    # Arkusz "Summary" jest szybkim widokiem projektu i rekomendacji.
    ws_summary = workbook.active
    ws_summary.title = "Summary"

    ws_summary.append(["Code Learning Analyzer v2"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])

    ws_summary.append(["Sciezka", str(path)])
    ws_summary.append(["Liczba plikow", summary.files_count])
    ws_summary.append(["Wszystkie linie", summary.total_lines])
    ws_summary.append(["Linie kodu", summary.code_lines])
    ws_summary.append(["Linie puste", summary.empty_lines])
    ws_summary.append(["Linie komentarzy", summary.comment_lines])
    ws_summary.append(["Wykryte funkcje", summary.function_count])
    ws_summary.append([f"Linie > {LONG_LINE_LIMIT} znakow", summary.long_lines])
    ws_summary.append(["Liczba TODO", summary.todo_count])

    ws_summary.append([])
    ws_summary.append(["Najwieksze pliki (top 5)"])
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
    for item in biggest:
        ws_summary.append([item.file, item.total_lines])

    ws_summary.append([])
    ws_summary.append(["Statystyki zaawansowane"])
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
    for file_stats in results:
        advanced = advanced_stats_map[file_stats.file]
        complexity_per_100 = (advanced.cyclomatic_complexity / max(file_stats.code_lines, 1)) * 100
        ws_summary.append([f"{file_stats.file}:", ""])
        ws_summary.append(["  Zlozonosc cyklomatyczna", advanced.cyclomatic_complexity])
        ws_summary.append(["  Zlozonosc / 100 linii kodu", f"{complexity_per_100:.1f}"])
        ws_summary.append(["  Gestosc kodu", f"{advanced.code_density:.2%}"])
        ws_summary.append(["  Wspolczynnik duplikacji", f"{advanced.duplication_score:.2%}"])
        ws_summary.append(["  Importy", advanced.import_count])
        ws_summary.append(["  Klasy", advanced.class_count])
        ws_summary.append(["  Petle", advanced.loop_count])
        ws_summary.append(["  If-y", advanced.if_count])
        ws_summary.append(["  Try-catch", advanced.try_catch_count])
        ws_summary.append([])
    ws_summary.append([])
    ws_summary.append(["Co cwiczyc dalej"])
    ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
    for idx, tip in enumerate(tips, start=1):
        ws_summary.append([idx, tip])

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 120

    # Arkusz "Files" trzyma metryki per plik, zeby latwo filtrowac/sortowac w Excelu.
    ws_files = workbook.create_sheet(title="Files")
    ws_files.append(
        [
            "File",
            "Total Lines",
            "Code Lines",
            "Empty Lines",
            "Comment Lines",
            "Functions",
            f"Long Lines > {LONG_LINE_LIMIT}",
            "TODO",
            "Cyclomatic Complexity",
            "Complexity per 100 LOC",
            "Code Density",
            "Duplication Score",
            "Imports",
            "Classes",
            "Loops",
            "If Statements",
            "Try-Catch Blocks",
        ]
    )
    for cell in ws_files[1]:
        cell.font = Font(bold=True)

    for item in results:
        advanced = advanced_stats_map[item.file]
        ws_files.append(
            [
                item.file,
                item.total_lines,
                item.code_lines,
                item.empty_lines,
                item.comment_lines,
                item.function_count,
                item.long_lines,
                item.todo_count,
                advanced.cyclomatic_complexity,
                f"{(advanced.cyclomatic_complexity / max(item.code_lines, 1)) * 100:.1f}",
                f"{advanced.code_density:.2%}",
                f"{advanced.duplication_score:.2%}",
                advanced.import_count,
                advanced.class_count,
                advanced.loop_count,
                advanced.if_count,
                advanced.try_catch_count,
            ]
        )

    ws_files.column_dimensions["A"].width = 90
    ws_files.column_dimensions["B"].width = 12
    ws_files.column_dimensions["C"].width = 12
    ws_files.column_dimensions["D"].width = 12
    ws_files.column_dimensions["E"].width = 14
    ws_files.column_dimensions["F"].width = 10
    ws_files.column_dimensions["G"].width = 16
    ws_files.column_dimensions["H"].width = 8
    ws_files.column_dimensions["I"].width = 20
    ws_files.column_dimensions["J"].width = 18
    ws_files.column_dimensions["K"].width = 12
    ws_files.column_dimensions["L"].width = 16
    ws_files.column_dimensions["M"].width = 8
    ws_files.column_dimensions["N"].width = 8
    ws_files.column_dimensions["O"].width = 8
    ws_files.column_dimensions["P"].width = 12
    ws_files.column_dimensions["Q"].width = 16

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)


def create_html_report(
    path: Path,
    results: list[FileStats],
    out_path: Path,
    *,
    top_languages_only: bool = False,
    max_languages: int = 12,
    max_quality_rows: int = 300,
    max_file_rows: int = 600,
) -> None:
    """Create an interactive HTML report with practical, learning-focused sections."""
    try:
        import matplotlib
        # "Agg" umozliwia eksport wykresow do obrazow bez zaleznosci od wyswietlania okna.
        matplotlib.use("Agg", force=True)
        import matplotlib.pyplot as plt
        from datetime import datetime
    except ImportError as exc:
        raise RuntimeError(
            "Brak biblioteki matplotlib. Zainstaluj: python3 -m pip install matplotlib"
        ) from exc

    from html import escape
    import base64
    import os
    import tempfile

    # Najpierw liczymy dane, potem budujemy wykresy i na koncu skladamy finalny HTML.
    summary = summarize(results)
    biggest = sorted(results, key=lambda item: item.total_lines, reverse=True)[:5]
    tips = learning_tips(summary, results)
    metric_docs = metric_explanations()
    advanced_stats_map = build_advanced_stats_map(results)
    language_rows = language_usage_stats(results)
    truncation_notes: list[str] = []

    # Safe mode ogranicza skale raportu, dzieki czemu HTML szybciej sie laduje na slabszym sprzecie.
    language_limit = 5 if top_languages_only else max(1, max_languages)
    if len(language_rows) > language_limit:
        # Nadmiar jezykow scalamy do "Inne", aby tabela i wykres pozostaly czytelne.
        kept_rows = language_rows[: language_limit - 1]
        other_rows = language_rows[language_limit - 1 :]
        other_files = sum(int(item["files"]) for item in other_rows)
        other_code = sum(int(item["code_lines"]) for item in other_rows)
        other_share = sum(float(item["code_share"]) for item in other_rows)
        kept_rows.append(
            {
                "language": "Inne",
                "files": other_files,
                "code_lines": other_code,
                "code_share": other_share,
            }
        )
        language_rows = kept_rows
        truncation_notes.append(
            f"Użycie języków zostało skrócone do {language_limit} pozycji (reszta zgrupowana jako 'Inne')."
        )

    def severity_label(score: float) -> tuple[str, str]:
        # To etykieta pomocnicza do priorytetyzacji pracy, nie absolutna "ocena jakosci" kodu.
        if score >= 45:
            return "Wysoki", "sev-high"
        if score >= 25:
            return "Sredni", "sev-medium"
        return "Niski", "sev-low"

    quality_rows: list[dict[str, str | float]] = []
    for file_stats in results:
        advanced = advanced_stats_map[file_stats.file]
        # max(..., 1) chroni przed dzieleniem przez zero, np. gdy plik ma same komentarze.
        code_lines = max(file_stats.code_lines, 1)
        complexity_per_100 = (advanced.cyclomatic_complexity / code_lines) * 100
        long_ratio = file_stats.long_lines / code_lines
        duplication_ratio = advanced.duplication_score
        # Heurystyczny score laczy kilka metryk w jeden ranking "od czego zaczac poprawki".
        score = (complexity_per_100 * 2.2) + (long_ratio * 100 * 1.4) + (duplication_ratio * 100 * 1.1)
        sev_text, sev_class = severity_label(score)
        quality_rows.append(
            {
                "file_name": Path(file_stats.file).name,
                "complexity_per_100": complexity_per_100,
                "duplication_pct": duplication_ratio * 100,
                "long_lines_pct": long_ratio * 100,
                "todos": file_stats.todo_count,
                "severity": sev_text,
                "severity_class": sev_class,
                "score": score,
            }
        )

    quality_rows.sort(key=lambda item: item["score"], reverse=True)
    if len(quality_rows) > max_quality_rows:
        # Przy duzych projektach sciecie tabeli mocno poprawia czas ladowania strony.
        quality_rows = quality_rows[:max_quality_rows]
        truncation_notes.append(
            f"Tabela jakości została skrócona do {max_quality_rows} plików dla wydajności."
        )

    # "Focus rows" to top 5 hotspotow, ktore pokazywane sa na wykresie i na poczatku analizy.
    focus_rows = quality_rows[:5]
    file_rows_source = sorted(results, key=lambda item: item.total_lines, reverse=True)
    if len(file_rows_source) > max_file_rows:
        # Szczegolowa tabela bywa najciezsza, dlatego tez ma osobny limit.
        file_rows_source = file_rows_source[:max_file_rows]
        truncation_notes.append(
            f"Tabela szczegółów plików została skrócona do {max_file_rows} największych plików."
        )

    # Wykresy zapisujemy tymczasowo jako PNG i osadzamy w HTML przez base64 (jeden samowystarczalny plik).
    temp_dir = tempfile.mkdtemp()
    try:
        # Cztery wykresy zapisujemy oddzielnie, bo pozniej osadzamy je jako base64 w jednym pliku HTML.
        fig, axis = plt.subplots(figsize=(8, 6))
        labels = ["Kod", "Komentarze", "Puste"]
        sizes = [summary.code_lines, summary.comment_lines, summary.empty_lines]
        colors = ["#3B82F6", "#0EA5A6", "#F59E0B"]
        axis.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors, startangle=90)
        axis.set_title("Podział linii kodu", fontsize=14, fontweight="bold")
        chart1_path = os.path.join(temp_dir, "chart1.png")
        plt.savefig(chart1_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, axis = plt.subplots(figsize=(10, 6))
        file_names = [Path(item.file).name for item in biggest]
        file_lines = [item.total_lines for item in biggest]
        bars = axis.bar(range(len(file_names)), file_lines, color="#2563EB")
        axis.set_title("Największe pliki", fontsize=14, fontweight="bold")
        axis.set_ylabel("Liczba linii")
        axis.set_xticks(range(len(file_names)))
        axis.set_xticklabels(file_names, rotation=45, ha="right")
        for bar in bars:
            height = bar.get_height()
            axis.text(bar.get_x() + bar.get_width() / 2.0, height + max(file_lines) * 0.02, f"{height}", ha="center", va="bottom", fontweight="bold")
        chart2_path = os.path.join(temp_dir, "chart2.png")
        plt.savefig(chart2_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, axis = plt.subplots(figsize=(10, 6))
        focus_names = [str(item["file_name"]) for item in focus_rows] if focus_rows else []
        focus_complexity = [float(item["complexity_per_100"]) for item in focus_rows] if focus_rows else []
        axis.bar(range(len(focus_names)), focus_complexity, color="#DC2626")
        axis.set_title("Złożoność / 100 linii (najtrudniejsze pliki)", fontsize=14, fontweight="bold")
        axis.set_ylabel("Punkty")
        axis.set_xticks(range(len(focus_names)))
        axis.set_xticklabels(focus_names, rotation=45, ha="right")
        chart3_path = os.path.join(temp_dir, "chart3.png")
        plt.savefig(chart3_path, dpi=150, bbox_inches="tight")
        plt.close()

        fig, axis = plt.subplots(figsize=(10, 6))
        language_labels = [str(item["language"]) for item in language_rows]
        language_sizes = [int(item["code_lines"]) for item in language_rows]
        if language_sizes:
            total_language_lines = sum(language_sizes) or 1
            wedges, _, _ = axis.pie(
                language_sizes,
                labels=None,
                autopct=lambda pct: f"{pct:.1f}%" if pct >= 5 else "",
                startangle=90,
                pctdistance=0.78,
                wedgeprops={"width": 0.45, "edgecolor": "white"},
            )
            legend_labels = [
                f"{label} ({size / total_language_lines:.1%})"
                for label, size in zip(language_labels, language_sizes)
            ]
            axis.legend(
                wedges,
                legend_labels,
                loc="center left",
                bbox_to_anchor=(1.02, 0.5),
                frameon=False,
                title="Języki",
            )
            axis.set_title("Udział języków (linie kodu)", fontsize=14, fontweight="bold")
        else:
            axis.text(0.5, 0.5, "Brak danych", ha="center", va="center", transform=axis.transAxes)
            axis.set_title("Udział języków (linie kodu)", fontsize=14, fontweight="bold")
        plt.tight_layout()
        chart4_path = os.path.join(temp_dir, "chart4.png")
        plt.savefig(chart4_path, dpi=150, bbox_inches="tight")
        plt.close()

        def image_to_base64(image_path: str) -> str:
            # Base64 pozwala osadzic obraz bez zewnetrznych plikow obok HTML.
            with open(image_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode("utf-8")

        chart1_b64 = image_to_base64(chart1_path)
        chart2_b64 = image_to_base64(chart2_path)
        chart3_b64 = image_to_base64(chart3_path)
        chart4_b64 = image_to_base64(chart4_path)

        summary_cards = f"""
        <div class="summary-grid">
            <div class="stat-card"><div class="stat-number">{summary.files_count}</div><div class="stat-label">Pliki</div></div>
            <div class="stat-card"><div class="stat-number">{summary.total_lines}</div><div class="stat-label">Wszystkie linie</div></div>
            <div class="stat-card"><div class="stat-number">{summary.code_lines}</div><div class="stat-label">Linie kodu</div></div>
            <div class="stat-card"><div class="stat-number">{summary.function_count}</div><div class="stat-label">Funkcje</div></div>
            <div class="stat-card"><div class="stat-number">{summary.long_lines}</div><div class="stat-label">Długie linie</div></div>
            <div class="stat-card"><div class="stat-number">{summary.todo_count}</div><div class="stat-label">TODO</div></div>
        </div>
        """

        # Gorna lista "Krok 1..3" ma prowadzic uzytkownika do pierwszych praktycznych poprawek.
        top_actions_html = "".join(
            f'<li><strong>Krok {idx}:</strong> {escape(tip)}</li>'
            for idx, tip in enumerate(tips[:3], start=1)
        ) or "<li>Brak pilnych rekomendacji.</li>"

        metric_help_html = "".join(
            f"<li>{escape(item)}</li>" for item in metric_docs
        )

        quality_rows_html = "".join(
            (
                "<tr>"
                f"<td><code>{escape(str(item['file_name']))}</code></td>"
                f"<td>{float(item['complexity_per_100']):.1f}</td>"
                f"<td>{float(item['duplication_pct']):.1f}%</td>"
                f"<td>{float(item['long_lines_pct']):.1f}%</td>"
                f"<td>{int(item['todos'])}</td>"
                f"<td><span class='sev {escape(str(item['severity_class']))}'>{escape(str(item['severity']))}</span></td>"
                "</tr>"
            )
            for item in quality_rows
        )
        # To jest tabela priorytetyzacji - pokazuje, od ktorych plikow najczesciej warto zaczac porzadki.

        # Osobna tabela "all files" daje pelny kontekst, nie tylko najgorsze przypadki.
        files_rows_html = "".join(
            (
                "<tr>"
                f"<td><code>{escape(Path(file_stats.file).name)}</code></td>"
                f"<td>{file_stats.total_lines}</td>"
                f"<td>{file_stats.code_lines}</td>"
                f"<td>{file_stats.comment_lines}</td>"
                f"<td>{file_stats.function_count}</td>"
                f"<td>{file_stats.long_lines}</td>"
                f"<td>{file_stats.todo_count}</td>"
                "</tr>"
            )
            for file_stats in file_rows_source
        )

        tips_html = "".join(
            f'<div class="tip"><strong>{idx}.</strong> {escape(tip)}</div>'
            for idx, tip in enumerate(tips, 1)
        )

        language_rows_html = "".join(
            (
                "<tr>"
                f"<td><strong>{escape(str(item['language']))}</strong></td>"
                f"<td>{int(item['files'])}</td>"
                f"<td>{int(item['code_lines'])}</td>"
                f"<td>{float(item['code_share']) * 100:.1f}%</td>"
                "</tr>"
            )
            for item in language_rows
        )
        # Gdy dane byly przyciete przez limity, pokazujemy to jawnie dla transparentnosci raportu.
        truncation_html = "".join(f"<li>{escape(note)}</li>" for note in truncation_notes)

        html_content = f"""
<!DOCTYPE html>
<html lang="pl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Analiza Kodu - {escape(path.name)}</title>
    <style>
        :root {{
            --bg: #f3f6fb;
            --card: #ffffff;
            --text: #1f2937;
            --muted: #6b7280;
            --accent: #1d4ed8;
            --line: #e5e7eb;
            --ok: #15803d;
            --warn: #b45309;
            --bad: #b91c1c;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, sans-serif;
            line-height: 1.55;
            color: var(--text);
            max-width: 1280px;
            margin: 0 auto;
            padding: 18px;
            background: radial-gradient(circle at top right, #dbeafe, var(--bg) 38%);
        }}
        .card {{
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: 12px;
            box-shadow: 0 3px 12px rgba(0,0,0,0.06);
            padding: 20px;
            margin-bottom: 20px;
        }}
        h1 {{
            margin-top: 0;
            color: #0f172a;
            border-bottom: 3px solid var(--accent);
            padding-bottom: 8px;
        }}
        h2 {{
            margin-top: 0;
            color: #0f172a;
        }}
        p.meta {{
            color: var(--muted);
            margin: 4px 0;
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 14px;
        }}
        .stat-card {{
            background: #f8fbff;
            border: 1px solid #dbeafe;
            border-radius: 10px;
            padding: 14px;
            text-align: center;
        }}
        .stat-number {{
            font-size: 1.9em;
            font-weight: 700;
            color: var(--accent);
        }}
        .stat-label {{
            color: var(--muted);
            font-size: 0.85em;
            text-transform: uppercase;
            letter-spacing: 0.8px;
        }}
        .charts-section {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
            gap: 16px;
        }}
        .chart-container img {{
            width: 100%;
            border-radius: 8px;
            border: 1px solid var(--line);
        }}
        .actions li {{
            margin-bottom: 8px;
        }}
        details {{
            border: 1px solid var(--line);
            border-radius: 8px;
            padding: 10px 12px;
            background: #f8fafc;
        }}
        details summary {{
            cursor: pointer;
            font-weight: 600;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid var(--line);
            font-size: 0.95em;
        }}
        th {{
            background: #f8fafc;
            position: sticky;
            top: 0;
            z-index: 1;
        }}
        .table-wrap {{
            max-height: 360px;
            overflow: auto;
            border: 1px solid var(--line);
            border-radius: 8px;
        }}
        .sev {{
            display: inline-block;
            padding: 3px 9px;
            border-radius: 999px;
            font-size: 0.82em;
            font-weight: 700;
        }}
        .sev-low {{ background: #dcfce7; color: var(--ok); }}
        .sev-medium {{ background: #fef3c7; color: var(--warn); }}
        .sev-high {{ background: #fee2e2; color: var(--bad); }}
        .tip {{
            background: #eef4ff;
            border-left: 4px solid var(--accent);
            padding: 12px;
            border-radius: 6px;
            margin-bottom: 8px;
        }}
        .footer {{
            text-align: center;
            color: var(--muted);
            margin-top: 24px;
            font-size: 0.9em;
        }}
    </style>
</head>
<body>
    <div class="card">
        <h1>Analiza Kodu: {escape(path.name)}</h1>
        <p class="meta"><strong>Data analizy:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p class="meta"><strong>Ścieżka:</strong> {escape(str(path))}</p>
    </div>

    <div class="card">
        <h2>Podsumowanie</h2>
        {summary_cards}
    </div>

    {"<div class='card'><h2>Tryb wydajności</h2><ul>" + truncation_html + "</ul></div>" if truncation_notes else ""}

    <div class="card">
        <h2>Co poprawić najpierw</h2>
        <ol class="actions">
            {top_actions_html}
        </ol>
        <details>
            <summary>Jak czytać metryki</summary>
            <ol>
                {metric_help_html}
            </ol>
        </details>
    </div>

    <div class="card">
        <h2>Wykresy</h2>
        <div class="charts-section">
            <div class="chart-container">
                <h3>Podział linii kodu</h3>
                <img src="data:image/png;base64,{chart1_b64}" alt="Podział linii kodu">
            </div>
            <div class="chart-container">
                <h3>Największe pliki</h3>
                <img src="data:image/png;base64,{chart2_b64}" alt="Największe pliki">
            </div>
            <div class="chart-container">
                <h3>Złożoność / 100 linii</h3>
                <img src="data:image/png;base64,{chart3_b64}" alt="Złożoność / 100 linii">
            </div>
            <div class="chart-container">
                <h3>Udział języków</h3>
                <img src="data:image/png;base64,{chart4_b64}" alt="Udział języków">
            </div>
        </div>
    </div>

    <div class="card">
        <h2>Użycie języków</h2>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Język</th>
                        <th>Pliki</th>
                        <th>Linie kodu</th>
                        <th>Udział kodu</th>
                    </tr>
                </thead>
                <tbody>
                    {language_rows_html}
                </tbody>
            </table>
        </div>
    </div>

    <div class="card">
        <h2>Jakość per plik</h2>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Plik</th>
                        <th>Złożoność / 100</th>
                        <th>Duplikacja</th>
                        <th>Długie linie</th>
                        <th>TODO</th>
                        <th>Priorytet</th>
                    </tr>
                </thead>
                <tbody>
                    {quality_rows_html}
                </tbody>
            </table>
        </div>
    </div>

    <div class="card">
        <h2>Szczegóły plików</h2>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Plik</th>
                        <th>Wszystkie linie</th>
                        <th>Linie kodu</th>
                        <th>Komentarze</th>
                        <th>Funkcje</th>
                        <th>Długie linie</th>
                        <th>TODO</th>
                    </tr>
                </thead>
                <tbody>
                    {files_rows_html}
                </tbody>
            </table>
        </div>
    </div>

    <div class="card">
        <h2>Porady i rekomendacje</h2>
        {tips_html}
    </div>

    <div class="footer">
        Wygenerowano za pomocą Code Learning Analyzer v2
    </div>
</body>
</html>
"""

        with open(out_path, "w", encoding="utf-8") as file_handle:
            file_handle.write(html_content)
    finally:
        import shutil

        shutil.rmtree(temp_dir)
